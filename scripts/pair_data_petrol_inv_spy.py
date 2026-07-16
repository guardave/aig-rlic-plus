#!/usr/bin/env python3
"""
Data Stage: Petroleum Inventory (WTTSTUS1) x SPY
================================================

Mode-3 Dana dispatch for pair_id petrol_inv_spy.

Deliverables:
  - data/petrol_inv_spy_monthly_{start}_{end}.parquet  (+ _latest alias)
  - data/petrol_inv_spy_daily_{start}_{end}.parquet    (+ _latest alias)
  - data/petrol_inv_spy_monthly_schema.json
  - data/petrol_inv_spy_daily_schema.json
  - data/data_dictionary_petrol_inv_spy_{DATE_TAG}.csv
  - data/summary_stats_petrol_inv_spy_{DATE_TAG}.csv
  - data/missing_value_report_petrol_inv_spy_{DATE_TAG}.md
  - results/petrol_inv_spy/stationarity_tests_{DATE_TAG}.csv
  - results/petrol_inv_spy/interpretation_metadata.json
  - data/manifest.json, data/display_name_registry.csv updates

WTTSTUS1 is weekly. Monthly analysis data use weekly-to-monthly aggregation
(calendar-month mean stock level). Daily data use real-time Last-Value-Carry-
Forward: a report-week Friday observation becomes available five calendar days
later (EIA Wednesday release for prior week), then carries until the next
release. The daily panel includes days_since_release.
"""

import csv
import json
import os
import shutil
import subprocess
import warnings
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd
from dotenv import load_dotenv

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=RuntimeWarning)

np.random.seed(42)

PAIR_ID = "petrol_inv_spy"
INDICATOR = "petrol_inv"
TARGET = "spy"
DATE_TAG = "20260617"
BASE_DIR = Path("/workspaces/aig-rlic-plus")
DATA_DIR = BASE_DIR / "data"
RESULTS_DIR = BASE_DIR / "results" / PAIR_ID
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

NOW_ISO = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def repo_rel(path: Path) -> str:
    if not path.is_absolute():
        return str(path)
    return str(path.relative_to(BASE_DIR))


def read_data_master_wttstus1() -> pd.Series:
    """Read the project-audited Data Master WTTSTUS1 sheet."""
    xlsx = DATA_DIR / "Data Master.xlsx"
    df = pd.read_excel(xlsx, sheet_name="WTTSTUS1")
    df["date"] = pd.to_datetime(df["date"])
    s = df.set_index("date")["WTTSTUS1"].astype(float).sort_index()
    s.name = "petrol_inv_kb"
    assert not s.index.duplicated().any(), "WTTSTUS1 duplicate dates"
    assert s.notna().all(), "WTTSTUS1 has missing values"
    return s


def verify_pre_master() -> str:
    """Phase 0 gate: WTTSTUS1 identity must match Pre-master dictionary."""
    from openpyxl import load_workbook

    xlsx = DATA_DIR / "Data Master.xlsx"
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Pre-master"]
    found = None
    for row in ws.iter_rows(min_row=1, max_row=3):
        for cell in row:
            if cell.value == "WTTSTUS1":
                found = cell.column
                break
        if found:
            break
    if not found:
        raise RuntimeError("Phase 0 gate failed: WTTSTUS1 not found in Pre-master")
    desc = ws.cell(row=2, column=found).value or ""
    required = ["Weekly U.S. Ending Stocks", "Crude Oil and Petroleum Products", "Thousand Barrels", "EIA"]
    missing = [x for x in required if x not in desc]
    if missing:
        raise RuntimeError(f"Phase 0 gate failed: Pre-master description missing {missing}: {desc!r}")
    return f"Pre-master column {found} ({ws.cell(row=1, column=found).coordinate}) confirms: {desc.strip()}"


def fetch_yahoo_close(ticker: str, start: str, end: str, name: str) -> pd.Series:
    import yfinance as yf

    df = yf.download(ticker, start=start, end=end, progress=False, auto_adjust=True)
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    s = df["Close"].astype(float).copy()
    s.name = name
    s.index = pd.to_datetime(s.index)
    if s.index.tz is not None:
        s.index = s.index.tz_localize(None)
    if s.empty:
        raise RuntimeError(f"Yahoo fetch failed for {ticker}")
    return s.sort_index()


def try_fetch_fred(series_id: str, start: str, end: str, name: str) -> pd.Series | None:
    """Optional controls. FRED rejects WTTSTUS1 today; controls may still work."""
    import urllib.parse
    import urllib.request

    load_dotenv()  # repo-root .env; hardcoded fallback removed (key rotated after leak)
    key = os.environ.get("FRED_API_KEY")
    if not key:
        raise SystemExit("FRED_API_KEY not set — copy .env.example to .env, or run setup.sh.")
    params = urllib.parse.urlencode({
        "series_id": series_id,
        "api_key": key,
        "file_type": "json",
        "observation_start": start,
        "observation_end": end,
    })
    url = f"https://api.stlouisfed.org/fred/series/observations?{params}"
    try:
        with urllib.request.urlopen(url, timeout=45) as r:
            obs = json.load(r)["observations"]
    except Exception as exc:
        print(f"  WARN: FRED fetch failed for {series_id}: {exc}")
        return None
    vals = {pd.Timestamp(o["date"]): float(o["value"]) for o in obs if o["value"] != "."}
    if not vals:
        return None
    s = pd.Series(vals, name=name).sort_index()
    return s


def rolling_z(x: pd.Series, window: int, min_periods: int) -> pd.Series:
    r = x.rolling(window, min_periods=min_periods)
    return (x - r.mean()) / r.std()


def add_petrol_transforms(df: pd.DataFrame, base_col: str, weekly_periods: bool = False) -> pd.DataFrame:
    p = df[base_col]
    yoy_lag = 52 if weekly_periods else 12
    mom_lag = 4 if weekly_periods else 1
    short_lag = 13 if weekly_periods else 3
    med_lag = 26 if weekly_periods else 6
    z_win = 260 if weekly_periods else 60
    z_min = 156 if weekly_periods else 36

    df["petrol_inv_pct_yoy"] = (p / p.shift(yoy_lag) - 1) * 100
    df["petrol_inv_pct_chg"] = (p / p.shift(mom_lag) - 1) * 100
    df["petrol_inv_3m_pct"] = (p / p.shift(short_lag) - 1) * 100
    df["petrol_inv_6m_pct"] = (p / p.shift(med_lag) - 1) * 100
    df["petrol_inv_ma12_kb"] = p.rolling(12 if not weekly_periods else 52, min_periods=10 if not weekly_periods else 40).mean()
    df["petrol_inv_dev_trend_pct"] = (p / df["petrol_inv_ma12_kb"] - 1) * 100
    df["petrol_inv_zscore_60m"] = rolling_z(p, z_win, z_min)
    df["petrol_inv_yoy_zscore_60m"] = rolling_z(df["petrol_inv_pct_yoy"], z_win, z_min)
    df["petrol_inv_accel_pct"] = df["petrol_inv_pct_chg"].diff()
    return df


def build_monthly(w: pd.Series, spy: pd.Series, vix: pd.Series, dgs10: pd.Series | None) -> pd.DataFrame:
    complete_end = (w.index.max().to_period("M") - 1).to_timestamp("M")
    idx = pd.date_range(w.index.min().to_period("M").to_timestamp("M"), complete_end, freq="ME")
    df = pd.DataFrame(index=idx)
    df.index.name = "date"
    # Stock series: calendar-month average of weekly ending stocks.
    df["petrol_inv_kb"] = w.resample("ME").mean().reindex(idx)
    df = add_petrol_transforms(df, "petrol_inv_kb", weekly_periods=False)
    df["spy"] = spy.resample("ME").last().reindex(idx)
    df["vix"] = vix.resample("ME").last().reindex(idx)
    if dgs10 is not None:
        df["dgs10"] = dgs10.resample("ME").last().reindex(idx)
    s = df["spy"]
    df["spy_ret"] = s.pct_change()
    df["spy_fwd_1m"] = s.shift(-1) / s - 1
    df["spy_fwd_3m"] = s.shift(-3) / s - 1
    df["spy_fwd_6m"] = s.shift(-6) / s - 1
    df["spy_fwd_12m"] = s.shift(-12) / s - 1
    return df


def build_daily_lvcf(w: pd.Series, spy: pd.Series, vix: pd.Series, dgs10: pd.Series | None) -> pd.DataFrame:
    releases = pd.DataFrame({
        "report_week_end": w.index,
        "release_date": w.index + pd.Timedelta(days=5),
        "petrol_inv_kb": w.values,
    }).sort_values("release_date")
    # Use SPY trading days from SPY inception through the last known release date.
    spy_cut = spy.loc[: releases["release_date"].max()]
    df = pd.DataFrame(index=spy_cut.index)
    df.index.name = "date"
    df["spy"] = spy_cut
    df["vix"] = vix.reindex(df.index)
    if dgs10 is not None:
        df["dgs10"] = dgs10.reindex(df.index).ffill()

    left = df.reset_index().rename(columns={"date": "trade_date"})
    merged = pd.merge_asof(
        left.sort_values("trade_date"),
        releases,
        left_on="trade_date",
        right_on="release_date",
        direction="backward",
    ).set_index("trade_date")
    merged.index.name = "date"
    merged["days_since_release"] = (merged.index - merged["release_date"]).dt.days.astype("float64")
    merged["report_week_end"] = pd.to_datetime(merged["report_week_end"])
    merged["release_date"] = pd.to_datetime(merged["release_date"])

    # Weekly transforms are computed on release events and then carried forward.
    r = releases.set_index("release_date")[["petrol_inv_kb"]].copy()
    r = add_petrol_transforms(r, "petrol_inv_kb", weekly_periods=True)
    for c in [x for x in r.columns if x != "petrol_inv_kb"]:
        merged[c] = pd.merge_asof(
            left[["trade_date"]].sort_values("trade_date"),
            r[[c]].reset_index().sort_values("release_date"),
            left_on="trade_date",
            right_on="release_date",
            direction="backward",
        )[c].to_numpy()

    merged["spy_ret"] = merged["spy"].pct_change()
    for h in [1, 5, 21, 63, 126, 252]:
        merged[f"spy_fwd_{h}d"] = merged["spy"].shift(-h) / merged["spy"] - 1
    return merged


def run_stationarity(monthly: pd.DataFrame, weekly: pd.Series) -> pd.DataFrame:
    from arch.unitroot import ADF, KPSS

    weekly_df = pd.DataFrame({"petrol_inv_weekly_kb": weekly})
    weekly_df = add_petrol_transforms(weekly_df, "petrol_inv_weekly_kb", weekly_periods=True)
    variables = {
        "monthly_level": monthly["petrol_inv_kb"],
        "monthly_delta": monthly["petrol_inv_kb"].diff(),
        "monthly_pct_chg": monthly["petrol_inv_pct_chg"],
        "monthly_yoy": monthly["petrol_inv_pct_yoy"],
        "monthly_zscore": monthly["petrol_inv_zscore_60m"],
        "monthly_yoy_zscore": monthly["petrol_inv_yoy_zscore_60m"],
        "daily_spy_ret": monthly["spy_ret"],
        "weekly_level": weekly_df["petrol_inv_weekly_kb"],
        "weekly_delta": weekly_df["petrol_inv_weekly_kb"].diff(),
        "weekly_pct_chg_4w": weekly_df["petrol_inv_pct_chg"],
        "weekly_yoy": weekly_df["petrol_inv_pct_yoy"],
        "weekly_zscore_5y": weekly_df["petrol_inv_zscore_60m"],
    }
    rows = []
    for name, s in variables.items():
        x = s.replace([np.inf, -np.inf], np.nan).dropna()
        for test in ["ADF", "KPSS"]:
            try:
                if test == "ADF":
                    obj = ADF(x, max_lags=12 if len(x) < 1000 else 52)
                    p = float(obj.pvalue)
                    rows.append({
                        "variable": name,
                        "test": test,
                        "statistic": round(float(obj.stat), 4),
                        "p_value": round(p, 4),
                        "lags": int(obj.lags),
                        "conclusion": "Stationary at 5%" if p < 0.05 else "Non-stationary",
                    })
                else:
                    obj = KPSS(x)
                    p = float(obj.pvalue)
                    rows.append({
                        "variable": name,
                        "test": test,
                        "statistic": round(float(obj.stat), 4),
                        "p_value": round(p, 4),
                        "lags": int(obj.lags),
                        "conclusion": "Fail to reject stationarity" if p > 0.05 else "Reject stationarity at 5%",
                    })
            except Exception as exc:
                rows.append({
                    "variable": name,
                    "test": test,
                    "statistic": np.nan,
                    "p_value": np.nan,
                    "lags": np.nan,
                    "conclusion": f"failed: {exc}",
                })
    return pd.DataFrame(rows)


def metadata_for_columns(df: pd.DataFrame, frequency: str, parquet_path: Path) -> dict:
    common = {
        "date": ("datetime64[ns]", "date", "Date", "neutral", f"{frequency.title()} date index."),
        "petrol_inv_kb": ("float64", "count", "Petroleum Product Stocks (kb)", "neutral",
                          "EIA WTTSTUS1: Weekly U.S. ending stocks of crude oil and petroleum products, thousand barrels. Monthly panel uses calendar-month mean; daily panel is release-lagged LVCF."),
        "petrol_inv_pct_yoy": ("float64", "pct", "Petroleum Stocks YoY (%)", "neutral",
                               "Percent change from one year earlier. Recommended stationary signal; captures inventory-cycle pressure while reducing secular trend."),
        "petrol_inv_pct_chg": ("float64", "pct", "Petroleum Stocks Short Change (%)", "neutral",
                               "Short-horizon percent change: monthly MoM in monthly panel; 4-week change in daily LVCF panel."),
        "petrol_inv_3m_pct": ("float64", "pct", "Petroleum Stocks 3M Change (%)", "neutral",
                              "Three-month/13-week percent change in petroleum stocks."),
        "petrol_inv_6m_pct": ("float64", "pct", "Petroleum Stocks 6M Change (%)", "neutral",
                              "Six-month/26-week percent change in petroleum stocks."),
        "petrol_inv_ma12_kb": ("float64", "count", "Petroleum Stocks 12M MA (kb)", "neutral",
                               "Rolling 12-month / 52-week average inventory level, thousand barrels."),
        "petrol_inv_dev_trend_pct": ("float64", "pct", "Petroleum Stocks vs Trend (%)", "neutral",
                                     "Percent deviation of petroleum stocks from 12-month / 52-week moving average."),
        "petrol_inv_zscore_60m": ("float64", "none", "Petroleum Stocks 5Y Z-Score", "neutral",
                                  "Rolling five-year z-score of inventory level. Seasonal and secular-trend contamination remains; prefer YoY or deviation-from-trend for stationary modeling."),
        "petrol_inv_yoy_zscore_60m": ("float64", "none", "Petroleum Stocks YoY 5Y Z-Score", "neutral",
                                      "Rolling five-year z-score of YoY percent change. Recommended stationary standardized transform."),
        "petrol_inv_accel_pct": ("float64", "pct", "Petroleum Stocks Acceleration (pp)", "neutral",
                                 "First difference of short-horizon percent change, in percentage points."),
        "spy": ("float64", "price", "SPY Close ($)", "neutral", "SPY adjusted close from Yahoo Finance."),
        "vix": ("float64", "index", "VIX Index Level", "lower_is_better", "CBOE VIX close from Yahoo Finance."),
        "dgs10": ("float64", "percent", "US 10Y Treasury Yield (%)", "neutral", "FRED DGS10, percent; daily panel forward-filled over market holidays."),
        "spy_ret": ("float64", "decimal_return", "SPY Daily Return (decimal)" if frequency == "daily" else "SPY Monthly Return", "neutral", "Simple SPY return over the panel frequency."),
        "spy_fwd_1m": ("float64", "decimal_return", "SPY Forward 1M Return", "neutral", "Forward one-month SPY return, decimal."),
        "spy_fwd_3m": ("float64", "decimal_return", "SPY Forward 3M Return", "neutral", "Forward three-month SPY return, decimal."),
        "spy_fwd_6m": ("float64", "decimal_return", "SPY Forward 6M Return", "neutral", "Forward six-month SPY return, decimal."),
        "spy_fwd_12m": ("float64", "decimal_return", "SPY Forward 12M Return", "neutral", "Forward twelve-month SPY return, decimal."),
        "report_week_end": ("datetime64[ns]", "date", "WTTSTUS1 Report Week End", "neutral", "Friday report-week date associated with the carried WTTSTUS1 value."),
        "release_date": ("datetime64[ns]", "date", "WTTSTUS1 Release Date", "neutral", "Assumed public availability date: report_week_end + 5 calendar days (EIA Wednesday release)."),
        "days_since_release": ("float64", "count", "Days Since Petroleum Release", "neutral", "Calendar days since the most recent WTTSTUS1 release available to the market."),
    }
    for h in [1, 5, 21, 63, 126, 252]:
        common[f"spy_fwd_{h}d"] = ("float64", "decimal_return", f"SPY {h}D Forward Return", "neutral", f"Forward {h}-trading-day SPY return, decimal.")

    cols = {
        "date": {
            "dtype": "datetime64[ns]",
            "unit": "date",
            "display_name": "Date",
            "direction": "neutral",
            "description": f"{frequency.title()} date index; spans {df.index.min().date()} through {df.index.max().date()}.",
        }
    }
    for c in df.columns:
        dtype, unit, display, direction, desc = common[c]
        cols[c] = {
            "dtype": str(df[c].dtype) if not pd.api.types.is_datetime64_any_dtype(df[c]) else "datetime64[ns]",
            "unit": unit,
            "display_name": display,
            "direction": direction,
            "description": desc,
            "source_reference": "Data Master.xlsx:WTTSTUS1 / EIA" if c.startswith("petrol") or c in {"report_week_end", "release_date", "days_since_release"} else ("yahoo:SPY" if c.startswith("spy") or c == "spy" else ("yahoo:^VIX" if c == "vix" else "FRED:DGS10")),
            "refresh_ttl_days": 7 if c.startswith("petrol") or c in {"report_week_end", "release_date", "days_since_release"} else 1,
        }
    return {
        "pair_id": PAIR_ID,
        "parquet_path": repo_rel(parquet_path),
        "schema_version": "1.0.0",
        "generated_at": NOW_ISO,
        "columns": cols,
    }


def write_data_dictionary(monthly: pd.DataFrame, daily: pd.DataFrame, path: Path) -> None:
    rows = []
    for frequency, df in [("monthly", monthly), ("daily", daily)]:
        meta = metadata_for_columns(df, frequency, Path(f"data/{PAIR_ID}_{frequency}_placeholder.parquet"))
        for c in ["date"] + list(df.columns):
            m = meta["columns"][c]
            rows.append({
                "frequency": frequency,
                "column_name": c,
                "display_name": m["display_name"],
                "description": m["description"],
                "source": "EIA via Data Master" if c.startswith("petrol") or c in {"report_week_end", "release_date", "days_since_release"} else ("Yahoo Finance" if c.startswith("spy") or c in {"spy", "vix"} else "FRED"),
                "series_id": "WTTSTUS1" if c.startswith("petrol") or c in {"report_week_end", "release_date", "days_since_release"} else ("SPY" if c.startswith("spy") or c == "spy" else ("^VIX" if c == "vix" else "DGS10")),
                "unit": m["unit"],
                "transformation": "LVCF from weekly release" if frequency == "daily" and c.startswith("petrol") else ("Weekly-to-monthly calendar mean" if frequency == "monthly" and c == "petrol_inv_kb" else "See description"),
                "seasonal_adj": "NSA (inventory stock; strong seasonality)" if c.startswith("petrol") else "N/A",
                "direction_convention": "Higher inventory = looser physical petroleum supply / possible demand weakness; sign is ambiguous for SPY and must be estimated.",
                "effective_start": str((df.index if c == "date" else df[c].dropna().index).min().date()) if (c == "date" or df[c].notna().any()) else "",
                "known_quirks": "Weekly stock data; EIA Wednesday release for prior-week data (~5-day lag from report week); strong refinery/driving-season seasonality and secular trend; FRED public API rejected WTTSTUS1 on 2026-06-17, so project-audited Data Master sheet was used.",
                "display_note": "Petroleum stocks are seasonal and slow-moving; daily values are step functions until the next weekly EIA release.",
                "refresh_freq": "weekly",
                "refresh_source": "EIA / FRED WTTSTUS1",
            })
    pd.DataFrame(rows).to_csv(path, index=False)


def update_display_registry(monthly: pd.DataFrame, daily: pd.DataFrame) -> None:
    reg_path = DATA_DIR / "display_name_registry.csv"
    reg = pd.read_csv(reg_path)
    existing = set(reg["column_name"])
    rows = []
    meta = {}
    for frequency, df in [("monthly", monthly), ("daily", daily)]:
        meta.update(metadata_for_columns(df, frequency, Path("x"))["columns"])
    for c, m in meta.items():
        if c in existing:
            continue
        rows.append({
            "column_name": c,
            "display_name": m["display_name"],
            "unit": m["unit"],
            "axis_label": m["display_name"],
        })
    if rows:
        reg = pd.concat([reg, pd.DataFrame(rows)], ignore_index=True)
        reg.to_csv(reg_path, index=False)

    reg_json_path = DATA_DIR / "display_name_registry.json"
    if reg_json_path.exists():
        with open(reg_json_path) as f:
            obj = json.load(f)
        if isinstance(obj, dict) and "entries" in obj:
            have = {e["column_name"] for e in obj["entries"]}
            obj["entries"].extend([r for r in rows if r["column_name"] not in have])
            obj["generated_at"] = NOW_ISO
            with open(reg_json_path, "w") as f:
                json.dump(obj, f, indent=2)


def update_manifest(monthly_path: Path, monthly_latest: Path, daily_path: Path, daily_latest: Path) -> None:
    manifest_path = DATA_DIR / "manifest.json"
    with open(manifest_path) as f:
        manifest = json.load(f)
    artifacts = [a for a in manifest["artifacts"] if PAIR_ID not in a.get("path", "")]
    base = {
        "source": "EIA/FRED:WTTSTUS1 via project Data Master.xlsx + yahoo:SPY/^VIX + optional FRED:DGS10",
        "refresh_ttl_days": 7,
        "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "pairs": [PAIR_ID],
        "mixed_freq_ttl_note": "Weekly petroleum inventory plus daily market data; TTL=7 days per fastest economically material indicator cadence. Daily SPY can refresh faster, but signal updates weekly.",
    }
    entries = [
        {**base, "path": repo_rel(monthly_path), "schema_ref": f"data/{PAIR_ID}_monthly_schema.json"},
        {**base, "path": repo_rel(monthly_latest), "source": f"alias_of:{repo_rel(monthly_path)}", "schema_ref": f"data/{PAIR_ID}_monthly_schema.json", "source_master": repo_rel(monthly_path)},
        {**base, "path": repo_rel(daily_path), "schema_ref": f"data/{PAIR_ID}_daily_schema.json"},
        {**base, "path": repo_rel(daily_latest), "source": f"alias_of:{repo_rel(daily_path)}", "schema_ref": f"data/{PAIR_ID}_daily_schema.json", "source_master": repo_rel(daily_path)},
    ]
    manifest["artifacts"] = artifacts + entries
    manifest["generated_at"] = NOW_ISO
    with open(manifest_path, "w") as f:
        json.dump(manifest, f, indent=2)


def update_prospective_pairs() -> None:
    path = DATA_DIR / "prospective_pairs.csv"
    with open(path, newline="") as f:
        rows = list(csv.DictReader(f))
        fields = f.readline()
    changed = False
    for row in rows:
        if row["pair_id"] == PAIR_ID:
            if row["status"] != "in_progress":
                row["status"] = "in_progress"
                changed = True
            break
    else:
        raise RuntimeError(f"{PAIR_ID} not found in prospective_pairs.csv")
    if changed:
        with open(path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)


def validate_json(schema: str, instance: str) -> None:
    cmd = ["python3", "scripts/validate_schema.py", "--schema", schema, "--instance", instance]
    subprocess.run(cmd, cwd=BASE_DIR, check=True)


def main() -> None:
    print("=" * 72)
    print("PHASE 0: DATA MASTER VERIFICATION")
    print("=" * 72)
    phase0 = verify_pre_master()
    print(phase0)

    print("\n" + "=" * 72)
    print("SOURCING")
    print("=" * 72)
    weekly = read_data_master_wttstus1()
    print(f"WTTSTUS1: {len(weekly)} obs, {weekly.index.min().date()} to {weekly.index.max().date()}")
    # FRED WTTSTUS1 public API check, documented but non-blocking because Data Master is authoritative here.
    fred_wtt = try_fetch_fred("WTTSTUS1", "1990-01-01", "2026-06-17", "fred_check")
    print("FRED WTTSTUS1 public API check:", "available" if fred_wtt is not None else "rejected/unavailable; using Data Master")

    end_for_market = (weekly.index.max() + pd.Timedelta(days=5) + pd.Timedelta(days=300)).strftime("%Y-%m-%d")
    spy = fetch_yahoo_close("SPY", "1993-01-01", end_for_market, "spy")
    vix = fetch_yahoo_close("^VIX", "1993-01-01", end_for_market, "vix")
    dgs10 = try_fetch_fred("DGS10", "1990-01-01", end_for_market, "dgs10")

    print("\n" + "=" * 72)
    print("BUILD PANELS")
    print("=" * 72)
    monthly = build_monthly(weekly, spy, vix, dgs10)
    daily = build_daily_lvcf(weekly, spy, vix, dgs10)
    print(f"Monthly: {monthly.shape}, {monthly.index.min().date()} to {monthly.index.max().date()}")
    print(f"Daily:   {daily.shape}, {daily.index.min().date()} to {daily.index.max().date()}")

    assert not monthly.index.duplicated().any()
    assert not daily.index.duplicated().any()
    assert monthly.index.is_monotonic_increasing
    assert daily.index.is_monotonic_increasing
    assert daily["days_since_release"].dropna().ge(0).all()
    assert daily["days_since_release"].dropna().max() <= 10, "weekly LVCF staleness unexpectedly exceeds 10 calendar days inside sample"
    assert monthly.index.max() == pd.Timestamp("2025-09-30"), "Monthly should stop at last complete month"
    assert daily.index.max() <= weekly.index.max() + pd.Timedelta(days=5), "Daily panel should not carry beyond last known release"

    print("\n" + "=" * 72)
    print("STATIONARITY")
    print("=" * 72)
    stat = run_stationarity(monthly, weekly)
    stat_path = RESULTS_DIR / f"stationarity_tests_{DATE_TAG}.csv"
    stat.to_csv(stat_path, index=False)
    print(stat.to_string(index=False))

    print("\n" + "=" * 72)
    print("WRITE ARTIFACTS")
    print("=" * 72)
    m_start, m_end = monthly.index.min().strftime("%Y%m%d"), monthly.index.max().strftime("%Y%m%d")
    d_start, d_end = daily.index.min().strftime("%Y%m%d"), daily.index.max().strftime("%Y%m%d")
    monthly_path = DATA_DIR / f"{PAIR_ID}_monthly_{m_start}_{m_end}.parquet"
    daily_path = DATA_DIR / f"{PAIR_ID}_daily_{d_start}_{d_end}.parquet"
    monthly_latest = DATA_DIR / f"{PAIR_ID}_monthly_latest.parquet"
    daily_latest = DATA_DIR / f"{PAIR_ID}_daily_latest.parquet"
    monthly.to_parquet(monthly_path)
    daily.to_parquet(daily_path)
    shutil.copy2(monthly_path, monthly_latest)
    shutil.copy2(daily_path, daily_latest)

    monthly_schema = metadata_for_columns(monthly, "monthly", monthly_path)
    daily_schema = metadata_for_columns(daily, "daily", daily_path)
    monthly_schema_path = DATA_DIR / f"{PAIR_ID}_monthly_schema.json"
    daily_schema_path = DATA_DIR / f"{PAIR_ID}_daily_schema.json"
    with open(monthly_schema_path, "w") as f:
        json.dump(monthly_schema, f, indent=2)
    with open(daily_schema_path, "w") as f:
        json.dump(daily_schema, f, indent=2)

    dict_path = DATA_DIR / f"data_dictionary_{PAIR_ID}_{DATE_TAG}.csv"
    write_data_dictionary(monthly, daily, dict_path)
    monthly.describe().T.round(4).to_csv(DATA_DIR / f"summary_stats_{PAIR_ID}_monthly_{DATE_TAG}.csv")
    daily.describe(include="all").T.to_csv(DATA_DIR / f"summary_stats_{PAIR_ID}_daily_{DATE_TAG}.csv")

    mv_path = DATA_DIR / f"missing_value_report_{PAIR_ID}_{DATE_TAG}.md"
    mv_lines = [
        f"# Missing Value Report - {PAIR_ID} ({DATE_TAG})",
        "",
        f"Monthly dataset: `{repo_rel(monthly_path)}` shape {monthly.shape}, {monthly.index.min().date()} to {monthly.index.max().date()}.",
        f"Daily dataset: `{repo_rel(daily_path)}` shape {daily.shape}, {daily.index.min().date()} to {daily.index.max().date()}.",
        "",
        "## Phase 0 / Source Check",
        "",
        phase0,
        "",
        "FRED public API rejected `WTTSTUS1` on 2026-06-17; the project-audited `data/Data Master.xlsx` sheet `WTTSTUS1` was used. Pre-master units/source agree with the dispatch brief and catalog.",
        "",
        "## Real-Time Lag",
        "",
        "Daily LVCF uses `release_date = report_week_end + 5 calendar days`, matching the EIA Wednesday release for prior-week data. Evan should not test leads shorter than this availability floor; use at least a 5-trading-day / one-week floor for daily models.",
        "",
        "## Missing Values",
        "",
        "| Dataset | Column | NaN count | Note |",
        "|---|---|---:|---|",
    ]
    for label, df in [("monthly", monthly), ("daily", daily)]:
        for c in df.columns:
            n = int(df[c].isna().sum())
            note = "leading transform / SPY pre-inception / forward-return tail" if n else "none"
            mv_lines.append(f"| {label} | `{c}` | {n} | {note} |")
    mv_lines.extend([
        "",
        "No internal gaps in WTTSTUS1. The daily indicator is an intentional step function; `days_since_release` documents staleness.",
        "",
        "## Stationarity Verdict",
        "",
        "Petroleum inventory levels are non-stationary and visibly seasonal/trending. Recommended signals for Evan: `petrol_inv_pct_yoy`, `petrol_inv_yoy_zscore_60m`, and `petrol_inv_dev_trend_pct`; avoid raw levels except as diagnostics.",
    ])
    mv_path.write_text("\n".join(mv_lines) + "\n")

    meta = {
        "pair_id": PAIR_ID,
        "schema_version": "1.1.0",
        "indicator": INDICATOR,
        "target": TARGET,
        "indicator_nature": "coincident",
        "indicator_type": "macro",
        "strategy_objective": "max_sharpe",
        "expected_direction": "mixed",
        "data_provenance": {
            "source": "EIA weekly petroleum stocks via project Data Master.xlsx (sheet WTTSTUS1); FRED public API rejected WTTSTUS1 on 2026-06-17",
            "series_id": "WTTSTUS1",
            "accessed_at": NOW_ISO,
        },
        "known_stress_episodes": [
            {"label": "GFC oil-demand collapse", "start": "2008-09-01", "end": "2009-06-30", "note": "Petroleum stocks rose sharply as demand weakened into recession."},
            {"label": "COVID demand shock", "start": "2020-03-01", "end": "2020-06-30", "note": "Inventories rose during the abrupt collapse in mobility and fuel demand."},
        ],
        "related_pair_ids": ["crude_oil_xle", "gold_copper_xli"],
        "owner_writes": {
            "dana": ["pair_id", "schema_version", "indicator", "target", "indicator_nature", "indicator_type", "data_provenance", "known_stress_episodes", "related_pair_ids"],
            "evan": ["observed_direction", "direction_consistent", "key_finding", "confidence"],
            "ray": ["strategy_objective", "expected_direction", "mechanism", "caveats", "narrative_summary"],
        },
        "last_updated_by": "dana",
        "last_updated_at": NOW_ISO,
    }
    meta_path = RESULTS_DIR / "interpretation_metadata.json"
    with open(meta_path, "w") as f:
        json.dump(meta, f, indent=2)

    update_display_registry(monthly, daily)
    update_manifest(monthly_path, monthly_latest, daily_path, daily_latest)
    update_prospective_pairs()

    print("\n" + "=" * 72)
    print("VALIDATE")
    print("=" * 72)
    validate_json("docs/schemas/data_subject.schema.json", repo_rel(monthly_schema_path))
    validate_json("docs/schemas/data_subject.schema.json", repo_rel(daily_schema_path))
    validate_json("docs/schemas/interpretation_metadata.schema.json", repo_rel(meta_path))
    validate_json("docs/schemas/data_manifest.schema.json", "data/manifest.json")

    print("\nDANA DONE")
    for p in [
        monthly_path, monthly_latest, daily_path, daily_latest, monthly_schema_path, daily_schema_path,
        dict_path, DATA_DIR / f"summary_stats_{PAIR_ID}_monthly_{DATE_TAG}.csv",
        DATA_DIR / f"summary_stats_{PAIR_ID}_daily_{DATE_TAG}.csv", mv_path, stat_path, meta_path,
        DATA_DIR / "manifest.json", DATA_DIR / "display_name_registry.csv", DATA_DIR / "prospective_pairs.csv",
    ]:
        print(f"- {repo_rel(p)}")


if __name__ == "__main__":
    main()
